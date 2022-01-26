from openpyxl import *
from win32com.client import Dispatch
import os

def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


class ExcelOp(object):
    def __init__(self, file, sheetIndex=0):
        self.file = file
        self.wb = load_workbook(self.file, data_only=True)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[sheetIndex]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def get_row_num(self):
        rows = self.ws.max_row
        return rows

    def get_col_num(self):
        columns = self.ws.max_column
        return columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column ).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data


    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


chineseDict = {
    'index': '序号',
    'name': '姓名',
    'salary': '工资',
    'rewardMoney': '奖励性绩效',
    'companySubsidy': '分公司补贴',
    'informalOverwork': '非编加班绩效',
    'yearMoney': '年终绩效奖',
    'performMoney': '文艺汇演奖励',
    'incomeSum': '小计',
    'accumulationFund': '公积金',
    'agedInsurance': '养老保险',
    'occupationYearMoney': '职业年金',
    'medicalInsurance': '医疗保险',
    'lossJobInsurance': '失业保险',
    'rentSubsidy': '提租补贴',
    'deductSum': '小计',
    'personalTax': '应纳个人所得税'
}


def initDict():
    return {
        # 'index': 0,
        'name': 0,
        'salary': 0,
        'rewardMoney': 0,
        'companySubsidy': 0,
        'informalOverwork': 0,
        'yearMoney': 0,
        'performMoney': 0,
        'incomeSum': 0,
        'accumulationFund': 0,
        'agedInsurance': 0,
        'occupationYearMoney': 0,
        'medicalInsurance': 0,
        'lossJobInsurance': 0,
        'rentSubsidy': 0,
        'deductSum': 0,
        'personalTax': 0
    }


def op_toExcel(data, fileName):  # openpyxl库储存数据到excel
    wb = Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表

    headList = []
    for k in chineseDict:
        headList.append(chineseDict[k])
    ws.append(headList)  # 添加表头

    dataList = []
    index = 1
    for i in data:
        rowList = [index]
        index += 1
        for j in data[i]:
            rowList.append(data[i][j])
        dataList.append(rowList)

    for i in range(len(dataList)):
        ws.append(dataList[i])

    wb.save(fileName)


if __name__ == "__main__":
    print('begin')
    fileName = input("请输入当前路径下的xlsx文件名，如abc.xlsx\n")
    formalSalary = ExcelOp(file=fileName, sheetIndex=0)
    temporarySalary = ExcelOp(file=fileName, sheetIndex=2)
    informalSalary = ExcelOp(file=fileName, sheetIndex=1)
    formalAdd = ExcelOp(file=fileName, sheetIndex=3)
    informalAdd = ExcelOp(file=fileName, sheetIndex=4)
    informalOverTime = ExcelOp(file=fileName, sheetIndex=11)

    formalSalary_beginRow = 0
    formalSalary_overRow = formalSalary.get_row_num()
    formalAdd_beginRow = 0
    formalAdd_overRow = formalAdd.get_row_num()
    informalSalary_beginRow = 0
    informalSalary_overRow = informalSalary.get_row_num()
    informalAdd_beginRow = 0
    informalAdd_overRow = formalAdd.get_row_num()
    informalOverTime_beginRow = 0
    informalOverTime_overRow = formalAdd.get_row_num()
    temporarySalary_beginRow = 0
    temporarySalary_overRow = temporarySalary.get_row_num()

    for i in range(formalSalary_overRow):
        val = formalSalary.get_cell_value(i + 1, 1)
        if val == 1:
            formalSalary_beginRow = i
            continue
        if val == '合计':
            formalSalary_overRow = i
            break

    for i in range(temporarySalary_overRow):
        val = temporarySalary.get_cell_value(i + 1, 1)
        if val == 1:
            temporarySalary_beginRow = i
            continue
        if val == '合计':
            temporarySalary_overRow = i
            break

    for i in range(informalSalary_overRow):
        val = informalSalary.get_cell_value(i + 1, 1)
        if val == 1:
            informalSalary_beginRow = i
            continue
        if val == '合计':
            informalSalary_overRow = i
            break

    for i in range(formalAdd_overRow):
        val = formalAdd.get_cell_value(i + 1, 1)
        if val == 1:
            formalAdd_beginRow = i
            continue
        if val == '合计':
            formalAdd_overRow = i
            break

    for i in range(informalAdd_overRow):
        val = informalAdd.get_cell_value(i + 1, 1)
        if val == 1:
            informalAdd_beginRow = i
            continue
        if val == '合计':
            informalAdd_overRow = i
            break

    for i in range(informalOverTime_overRow):
        val = informalOverTime.get_cell_value(i + 1, 1)
        if val == 1:
            informalOverTime_beginRow = i
            continue
        if val == '合计':
            informalOverTime_overRow = i
            break


    dataDict = dict()
    index = 0

    for i in range(formalSalary_beginRow, formalSalary_overRow):
        index += 1
        name = formalSalary.get_cell_value(i + 1, 2)
        if name not in dataDict:
            dataDict[name] = initDict()
        salary = formalSalary.get_cell_value(i + 1, 11) or 0

        accumulationFund = formalSalary.get_cell_value(i + 1, 14) or 0
        rentSubsidy = formalSalary.get_cell_value(i + 1, 8) or 0
        agedInsurance = formalSalary.get_cell_value(i + 1, 16) or 0
        occupationYearMoney = formalSalary.get_cell_value(i + 1, 17) or 0
        medicalInsurance = formalSalary.get_cell_value(i + 1, 18) or 0
        lossJobInsurance = formalSalary.get_cell_value(i + 1, 19) or 0

        dataDict[name].update({
            'name': formalSalary.get_cell_value(i + 1, 2),
            'salary': salary,
            'companySubsidy': 0,
            'informalOverwork': 0,
            'yearMoney': 0,
            'performMoney': 0,
            'incomeSum': '=SUM(C' + str(index + 1 ) + ':H' + str(index  + 1) + ')',
            'accumulationFund': accumulationFund,
            'agedInsurance': agedInsurance,
            'occupationYearMoney': occupationYearMoney,
            'medicalInsurance': medicalInsurance,
            'lossJobInsurance': lossJobInsurance,
            'rentSubsidy': rentSubsidy,
            'deductSum': '=SUM(J' + str(index + 1) + ':O' + str(index + 1) + ')',
            # 'deductSum': round(
            #     accumulationFund + agedInsurance + occupationYearMoney + medicalInsurance + lossJobInsurance + rentSubsidy,
            #     2),
            'personalTax': 0
        })

    for i in range(informalSalary_beginRow, informalSalary_overRow):
        index +=1
        name = informalSalary.get_cell_value(i + 1, 2)
        if name not in dataDict:
            dataDict[name] = initDict()

        salary = informalSalary.get_cell_value(i + 1, 3) or 0
        accumulationFund = informalSalary.get_cell_value(i + 1, 7) or 0
        agedInsurance = informalSalary.get_cell_value(i + 1, 9) or 0
        medicalInsurance = informalSalary.get_cell_value(i + 1, 11) or 0
        lossJobInsurance = informalSalary.get_cell_value(i + 1, 12) or 0

        dataDict[name].update({
            'name': informalSalary.get_cell_value(i + 1, 2),
            'salary': salary,
            'companySubsidy': 0,
            'yearMoney': 0,
            'performMoney': 0,
            'incomeSum': '=SUM(C' + str(index + 1) + ':H' + str(index + 1) + ')',
            'accumulationFund': accumulationFund,
            'agedInsurance': agedInsurance,
            'occupationYearMoney': 0,
            'medicalInsurance': medicalInsurance,
            'lossJobInsurance': lossJobInsurance,
            'rentSubsidy': 0,
            'deductSum': '=SUM(J' + str(index + 1) + ':O' + str(index + 1) + ')',
            # 'deductSum': round(
            #     accumulationFund + agedInsurance + medicalInsurance + lossJobInsurance,
            #     2),
            'personalTax': 0
        })

    for i in range(temporarySalary_beginRow, temporarySalary_overRow):
        index += 1
        name = temporarySalary.get_cell_value(i + 1, 2)
        if name not in dataDict:
            dataDict[name] = initDict()
        salary = temporarySalary.get_cell_value(i + 1, 3) or 0
        medicalSupply = temporarySalary.get_cell_value(i + 1, 9) or 0
        supplyMoney = temporarySalary.get_cell_value(i + 1, 10) or 0
        overWork = temporarySalary.get_cell_value(i + 1, 11) or 0
        reward = temporarySalary.get_cell_value(i + 1, 13) or 0

        rewardMoney = medicalSupply + supplyMoney + overWork + reward

        dataDict[name].update({
            'name': temporarySalary.get_cell_value(i + 1, 2),
            'salary': salary,
            'rewardMoney': round(rewardMoney, 2),
            'companySubsidy': 0,
            # 'informalOverwork': informalOverwork,
            'yearMoney': 0,
            'performMoney': 0,
            'incomeSum': '=SUM(C' + str(index + 1) + ':H' + str(index + 1) + ')',
            'accumulationFund': 0,
            'agedInsurance': 0,
            'occupationYearMoney': 0,
            'medicalInsurance': 0,
            'lossJobInsurance': 0,
            'rentSubsidy': 0,
            'deductSum': 0,
            'personalTax': 0
        })

    for i in range(formalAdd_beginRow, formalAdd_overRow):
        name = formalAdd.get_cell_value(i + 1, 2) or 0
        if name not in dataDict:
            dataDict[name] = initDict()
        dataDict[name].update({'rewardMoney': formalAdd.get_cell_value(i + 1, 8) or 0})

    for i in range(informalAdd_beginRow, informalAdd_overRow):
        name = informalAdd.get_cell_value(i + 1, 2) or 0
        if name not in dataDict:
            dataDict[name] = initDict()
        dataDict[name].update({'rewardMoney': informalAdd.get_cell_value(i + 1, 9) or 0})

    # for i in range(informalOverTime_beginRow, informalOverTime_overRow):
    #     name = informalOverTime.get_cell_value(i + 1, 2) or 0
    #     if name not in dataDict:
    #         dataDict[name] = initDict()
    #     dataDict[name].update({'informalOverwork': informalOverTime.get_cell_value(i + 1, 3) or 0})

    outputName = '输出文件.xlsx'
    op_toExcel(dataDict, outputName)

    just_open(os.getcwd() + "\\" + outputName)

    excel_op = ExcelOp(file=outputName)

    [row, col] = excel_op.get_row_clo_num()
    for i in range(3, col + 1):
        colValueArr = excel_op.get_col_value(i)
        total = 0
        for j in range(1, len(colValueArr)):
            if colValueArr[j] is not None:
                total += colValueArr[j]
        excel_op.set_cell_value(row + 1, i, total)
